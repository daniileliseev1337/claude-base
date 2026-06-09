# -*- coding: utf-8 -*-
"""
id-cascade tool (заготовка для skill id-cascade). ОБЕЗЛИЧЕНО / параметризовано.

Назначение: добавить набор позиций в КОНЕЦ блока каждого помещения в листах ИД-книги
(ПВ/ВОР), сохранив merge-заголовки и сквозную нумерацию «по прогонам».

Дисциплина (грабли, проверено на проде):
- merge снять ДО вставок -> вставка СНИЗУ-ВВЕРХ -> восстановить merge со сдвигом;
- перенумерация колонки № ПО ПРОГОНАМ (reset на строке-заголовке таблицы B=="Наименование"),
  а не сквозняком — чистит и предсуществующие дубли/дыры;
- стиль новых ячеек копируется от соседнего прибора (последняя строка блока);
- №Акта/Дата/Ед.изм — от соседа (не выдумывать);
- идемпотентность: если добавляемая позиция уже в блоке — abort;
- НЕ называть этот файл inspect.py (конфликт со stdlib inspect при import openpyxl);
- после прогона: открыть книгу в Excel и пересохранить (openpyxl не кэширует значения формул).

CONFIG ниже — единственное, что меняется под задачу. Пути хардкодить (UTF-8), вывод в файл.
"""
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter

# ====================== CONFIG (заполнить под задачу) ======================
WORKBOOK = r"<АБСОЛЮТНЫЙ ПУТЬ К .xlsx>"

# Позиции для вставки: список (наименование, {помещение: кол-во})
# Кол-во берётся из ЭТАЛОНА (ВСО as-built), по помещениям.
ITEMS = [
    ("<Наименование позиции 1>", {
        "Пом. A": 2, "Пом. B": 1,
    }),
    ("<Наименование позиции 2>", {
        "Пом. A": 4, "Пом. B": 2,
    }),
]

# Листы: имя -> схема колонок.
#   "ncols": число колонок; "name": колонка наименования; "num": колонка №;
#   "qty": колонка объёма; "copy_from_neighbor": список колонок, чьи значения копировать у соседа
#   (напр. №Акта, Дата, Ед.изм для ПВ); "const": {колонка: значение} фиксированные (напр. Ед.изм="шт" в ВОР).
SHEETS = {
    "<Лист ПВ>":  {"ncols": 7, "num": 1, "name": 2, "qty": 6, "copy_from_neighbor": [3, 4, 5], "const": {}},
    "<Лист ВОР>": {"ncols": 4, "num": 1, "name": 2, "qty": 4, "copy_from_neighbor": [],        "const": {3: "шт"}},
}
HEADER_NAME = "Наименование"   # текст в колонке наименования у строки-шапки таблицы (reset нумерации)
# ===========================================================================


def norm(s):
    return " ".join(str(s).split()) if s is not None else ""


def is_blank(ws, r, col):
    v = ws.cell(row=r, column=col).value
    return v is None or str(v).strip() == ""


def process_sheet(wb, sheet, cfg):
    ws = wb[sheet]
    nc, cnum, cname, cqty = cfg["ncols"], cfg["num"], cfg["name"], cfg["qty"]
    print("\n=== %r было max_row=%d merges=%d ===" % (sheet, ws.max_row, len(ws.merged_cells.ranges)))

    # все помещения, упомянутые в ITEMS
    poms = set()
    for _, by_pom in ITEMS:
        poms.update(by_pom.keys())
    poms_norm = {norm(p): p for p in poms}

    # найти строки-заголовки помещений + границу блока (первая строка с пустым наименованием ниже)
    blocks = {}
    for r in range(1, ws.max_row + 1):
        a = norm(ws.cell(row=r, column=cnum).value)
        if a in poms_norm:
            b = r + 1
            while b <= ws.max_row and not is_blank(ws, b, cname):
                b += 1
            blocks[poms_norm[a]] = (r, b)
    missing = [p for p in poms if p not in blocks]
    assert not missing, "В %r не найдены помещения: %s" % (sheet, missing)

    # идемпотентность
    add_names = {norm(n) for n, _ in ITEMS}
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(row=r, column=cname).value) in add_names:
            raise SystemExit("ОТМЕНА: позиция уже есть в %r (стр.%d)" % (sheet, r))

    # снять merge
    merges = [tuple(m.bounds) for m in ws.merged_cells.ranges]  # (min_col,min_row,max_col,max_row)
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    # шаблоны стиля + значения-соседи (до вставок)
    tmpl = {}
    for pom, (hr, br) in blocks.items():
        last = br - 1
        tmpl[pom] = {
            "fonts":  {c: copy(ws.cell(row=last, column=c).font) for c in range(1, nc+1)},
            "bords":  {c: copy(ws.cell(row=last, column=c).border) for c in range(1, nc+1)},
            "fills":  {c: copy(ws.cell(row=last, column=c).fill) for c in range(1, nc+1)},
            "aligns": {c: copy(ws.cell(row=last, column=c).alignment) for c in range(1, nc+1)},
            "nfmt":   {c: ws.cell(row=last, column=c).number_format for c in range(1, nc+1)},
            "neighbor": {c: ws.cell(row=last, column=c).value for c in cfg["copy_from_neighbor"]},
        }

    boundaries = [br for (_, br) in blocks.values()]

    def fill(pom, row, name, qty):
        ws.cell(row=row, column=cname).value = name
        ws.cell(row=row, column=cqty).value = qty
        for c, v in cfg["const"].items():
            ws.cell(row=row, column=c).value = v
        for c in cfg["copy_from_neighbor"]:
            ws.cell(row=row, column=c).value = tmpl[pom]["neighbor"][c]
        for c in range(1, nc+1):
            cell = ws.cell(row=row, column=c)
            cell.font = copy(tmpl[pom]["fonts"][c]); cell.border = copy(tmpl[pom]["bords"][c])
            cell.fill = copy(tmpl[pom]["fills"][c]); cell.alignment = copy(tmpl[pom]["aligns"][c])
            cell.number_format = tmpl[pom]["nfmt"][c]

    # вставки снизу-вверх: для каждого помещения столько строк, сколько позиций его касается
    plan = []
    for pom, (hr, br) in blocks.items():
        rows = [(n, by[pom]) for n, by in ITEMS if pom in by]
        plan.append((pom, br, rows))
    for pom, br, rows in sorted(plan, key=lambda x: x[1], reverse=True):
        ws.insert_rows(br, amount=len(rows))
        for i, (name, qty) in enumerate(rows):
            fill(pom, br + i, name, qty)

    # восстановить merge со сдвигом
    def shift(r):
        return sum(len([1 for n, by in ITEMS if pom in by]) for pom, b in
                   [(p, bb) for p, (hh, bb) in blocks.items()] if b <= r)
    for (c1, r1, c2, r2) in merges:
        ws.merge_cells(start_row=r1 + shift(r1), start_column=c1,
                       end_row=r2 + shift(r2), end_column=c2)

    # перенумерация по прогонам
    counter = 0; started = False
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=cname).value
        bs = "" if b is None else str(b).strip()
        if bs == "":
            continue
        if bs == HEADER_NAME:
            counter = 0; started = True; continue
        if not started:
            continue
        counter += 1
        ws.cell(row=r, column=cnum).value = counter

    print("  -> стало max_row=%d merges=%d" % (ws.max_row, len(ws.merged_cells.ranges)))


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False)
    for sheet, cfg in SHEETS.items():
        process_sheet(wb, sheet, cfg)
    wb.save(WORKBOOK)
    print("\nСОХРАНЕНО:", WORKBOOK, "\n⚠ открыть в Excel и пересохранить (кэш формул).")


if __name__ == "__main__":
    main()
