#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verify_document_entry.py — контрольная сумма документа (детерминированный
верификатор внесения первички: УПД / накладная / счёт / спецификация).

Ловит ВЕСЬ класс ошибок маппинга одной проверкой: пропуск позиции, дубль,
ошибка цены/кол-ва. Источник правила — feedback document-control-sum-rule
(реальный кейс: пропущена 21-я позиция на 4-м листе многостраничной накладной,
усугублено двойником по габариту; контрольная сумма поймала бы мгновенно).

Идея: после внесения N позиций документа в реестр —
    Σ(внесённые суммы по колонке) == «Всего к оплате» документа   И
    число непустых внесённых строк == число позиций документа.
Расхождение → почти наверняка пропуск/дубль/ошибка цены.

Зависит только от openpyxl (не от MCP/uvx-кэша).

Пример:
    python verify_document_entry.py reestr.xlsx --sheet "Спец. 20" \\
        --col O --rows 244:250 --expected-sum 923750.40 --expected-count 7
    # допуск суммы по умолчанию 0.01 (копейка); --tol для своего

Exit code: 0 = PASS, 1 = FAIL, 2 = ошибка аргументов/файла.
"""
import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("ОШИБКА: нужен openpyxl (pip install openpyxl)", file=sys.stderr)
    sys.exit(2)


def parse_rows(spec: str):
    """'244:250' -> (244, 250); '244' -> (244, 244)."""
    spec = spec.strip()
    if ":" in spec:
        a, b = spec.split(":", 1)
        return int(a), int(b)
    return int(spec), int(spec)


def as_number(v):
    """Привести значение ячейки к float или None (формулы без кэша → None)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        if not s or s.startswith("="):
            return None  # формула как текст / пусто — нет кэша значения
        try:
            return float(s)
        except ValueError:
            return None
    return None


def main():
    ap = argparse.ArgumentParser(description="Контрольная сумма документа при внесении первички")
    ap.add_argument("xlsx", help="путь к реестру .xlsx")
    ap.add_argument("--sheet", required=True, help="имя листа")
    ap.add_argument("--col", required=True, help="буква колонки суммы (например O)")
    ap.add_argument("--rows", required=True, help="диапазон строк, напр. 244:250")
    ap.add_argument("--expected-sum", type=float, required=True,
                    help="эталон: «Всего к оплате» документа")
    ap.add_argument("--expected-count", type=int, default=None,
                    help="эталон: число позиций документа (опционально)")
    ap.add_argument("--tol", type=float, default=0.01, help="допуск по сумме (коп.), default 0.01")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"ОШИБКА: файл не найден: {path}", file=sys.stderr)
        sys.exit(2)

    # data_only=True — кэш значений; формулы без кэша вернут None (это поймаем как пустую)
    wb = load_workbook(str(path), data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"ОШИБКА: лист '{args.sheet}' не найден. Есть: {wb.sheetnames}", file=sys.stderr)
        sys.exit(2)
    ws = wb[args.sheet]

    col = column_index_from_string(args.col.strip().upper())
    r1, r2 = parse_rows(args.rows)

    total = 0.0
    nonempty = 0
    empty_rows = []
    none_cache_rows = []
    for r in range(r1, r2 + 1):
        raw = ws.cell(row=r, column=col).value
        num = as_number(raw)
        if num is None:
            if raw is None:
                empty_rows.append(r)
            else:
                # есть содержимое, но не число (вероятно формула без кэша)
                none_cache_rows.append(r)
        else:
            total += num
            nonempty += 1

    diff = round(total - args.expected_sum, 2)
    sum_ok = abs(diff) <= args.tol
    count_ok = (args.expected_count is None) or (nonempty == args.expected_count)

    print(f"Лист: {args.sheet} | колонка {args.col.upper()} | строки {r1}:{r2}")
    print(f"Σ внесённых: {total:.2f}")
    print(f"Эталон (Всего к оплате): {args.expected_sum:.2f}")
    print(f"Расхождение: {diff:+.2f} (допуск ±{args.tol})  -> {'OK' if sum_ok else 'РАСХОЖДЕНИЕ'}")
    print(f"Непустых позиций: {nonempty}" +
          (f" | ожидалось {args.expected_count} -> {'OK' if count_ok else 'НЕ СОВПАЛО'}"
           if args.expected_count is not None else ""))
    if empty_rows:
        print(f"⚠ Пустые строки в диапазоне: {empty_rows} — возможна ПРОПУЩЕННАЯ позиция.")
    if none_cache_rows:
        print(f"⚠ Строки с формулой без кэша: {none_cache_rows} — открой/сохрани файл в Excel "
              f"для пересчёта, иначе сумма занижена.")

    if sum_ok and count_ok and not none_cache_rows:
        print("\nPASS — контрольная сумма сошлась.")
        sys.exit(0)

    print("\nFAIL — проверь маппинг:")
    if not sum_ok:
        if diff < 0:
            print(f"  • Σ меньше эталона на {abs(diff):.2f} — вероятно ПРОПУЩЕНА позиция "
                  f"(частый кейс: одиночная позиция на последнем листе многостраничного документа) "
                  f"или занижены цена/кол-во.")
        else:
            print(f"  • Σ больше эталона на {diff:.2f} — вероятно ДУБЛЬ позиции "
                  f"или завышены цена/кол-во.")
    if not count_ok:
        print(f"  • Число позиций {nonempty} ≠ ожидаемых {args.expected_count} "
              f"(проверь двойники по габариту: одинаковый размер, разное исполнение/цена).")
    if none_cache_rows:
        print("  • Есть формулы без кэша — сумма может быть неполной, пересчитай в Excel.")
    sys.exit(1)


if __name__ == "__main__":
    main()
