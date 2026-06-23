# -*- coding: utf-8 -*-
"""
co_verify — сверить таблицу СО из проектного PDF с данными ВОР/спецификации.
Детерминированно ловит: расхождение количеств, единиц, лишние/недостающие позиции.

Использование:
  python co_verify.py --pdf <СО.pdf> --json <data.json>
  python co_verify.py --pdf <СО.pdf> --xlsx <ВОР.xlsx> [--sheet Рабочий] [--cols A,B,F,G]
     (--cols = столбцы pos,name,unit,qty; по умолчанию A,B,F,G — шаблон ВОР)

Источник-json: {"rows":[{"pos","name","qty","unit","group"?}, ...]} ИЛИ плоский список тех же dict.
Позиции вида "СО1-1.2" трактуются как (co=1, pos=1.2). group=true — пропускается (заголовки).
Возвращает список расхождений и код выхода 0 (чисто) / 1 (есть расхождения).
"""
import sys, os, re, json, argparse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import co_engine as E

def split_pos(praw):
    m = re.match(r'^СО(\d+)-(.+)$', str(praw))
    return (m.group(1), m.group(2)) if m else ('', str(praw))

def load_json(path):
    d = json.load(open(path, encoding="utf-8"))
    rows = d["rows"] if isinstance(d, dict) and "rows" in d else d
    out = []
    for r in rows:
        if r.get("group"):
            continue
        co, pos = split_pos(r.get("pos", ""))
        out.append({'co': co, 'pos': pos, 'name': r.get('name', ''),
                    'unit': r.get('unit', ''), 'qty': E.norm_qty(r.get('qty'))})
    return out

def load_xlsx(path, sheet, cols):
    import openpyxl
    from openpyxl.utils import column_index_from_string as cix
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    cp, cn, cu, cq = [cix(c.strip()) - 1 for c in cols.split(',')]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if cp >= len(row) or row[cp] in (None, ""):
            continue
        co, pos = split_pos(row[cp])
        out.append({'co': co, 'pos': pos, 'name': row[cn] if cn < len(row) else '',
                    'unit': row[cu] if cu < len(row) else '', 'qty': E.norm_qty(row[cq] if cq < len(row) else None)})
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True)
    ap.add_argument("--json")
    ap.add_argument("--xlsx")
    ap.add_argument("--sheet", default="Рабочий")
    ap.add_argument("--cols", default="A,B,F,G")
    a = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    pdf_rows = E.extract_co(a.pdf)
    if a.json:
        data = load_json(a.json)
    elif a.xlsx:
        data = load_xlsx(a.xlsx, a.sheet, a.cols)
    else:
        print("Укажите --json или --xlsx"); sys.exit(2)

    pairs, only_data, only_pdf = E.pair_by_position_then_name(data, pdf_rows)
    issues = []
    for j, p in pairs:
        t = E.tag(j['co'], j['pos'])
        if E.norm_qty(j['qty']) != E.norm_qty(p['qty']):
            issues.append(("QTY", t, f"данные={j['qty']}  PDF={p['qty']} ({p['unit']}, стр.{p['page']}) «{p['name'][:42]}»"))
        if E.norm_unit(j['unit']) != E.norm_unit(p['unit']):
            issues.append(("UNIT", t, f"данные='{j['unit']}'  PDF='{p['unit']}' (стр.{p['page']}) «{p['name'][:30]}»"))
    for j in only_data:
        issues.append(("EXTRA_IN_DATA", E.tag(j['co'], j['pos']), f"в данных «{j['name'][:50]}» ({j['qty']} {j['unit']}), нет в PDF"))
    for p in only_pdf:
        issues.append(("MISSING_IN_DATA", E.tag(p['co'], p['pos']), f"в PDF «{p['name'][:50]}» ({p['qty']} {p['unit']}, стр.{p['page']}), нет в данных"))

    print(f"PDF позиций={len(pdf_rows)}  данные={len(data)}  пар={len(pairs)}  расхождений={len(issues)}")
    for typ, t, msg in sorted(issues, key=lambda x: (x[0], x[1])):
        print(f"  [{typ:16}] {t:12} {msg}")
    sys.exit(1 if issues else 0)

if __name__ == "__main__":
    main()
