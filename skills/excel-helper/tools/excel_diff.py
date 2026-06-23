#!/usr/bin/env python3
"""excel-helper Layer-3 — стабильные детерминированные функции для xlsx.

Вынесено из SKILL.md (skill-development правило 2: повторяемая логика → tools/),
чтобы не переписывать заново каждый раз (риск ошибки при ручном воспроизведении).

Функции (importable):
  cell_diff(v1, v2, out)        — cell-by-cell diff с подсветкой, новый xlsx
  formula_diff(v1, v2)          — расхождения формул как текст (не значений)
  find_formula_errors(path)     — найти #REF!/#DIV/0!/#NAME? (по cached-значениям)
  find_duplicates(path, key)    — дубликаты по колонке-ключу (pandas)

CLI:
  python excel_diff.py celldiff  v1.xlsx v2.xlsx out.xlsx
  python excel_diff.py formuladiff v1.xlsx v2.xlsx
  python excel_diff.py errors    file.xlsx
  python excel_diff.py dupes     file.xlsx ID

Ловушки (см. SKILL.md): data_only=True вернёт None для формул без кэша Excel;
openpyxl save() стирает drawing-слой (картинки) и ломает vertical-merge при delete_rows.
Эти функции — read/diff (errors, formuladiff, dupes безопасны); cell_diff пишет НОВЫЙ
файл из v2, для файлов с картинками проверь media-слой перед использованием результата.
"""
import sys
import io

# UTF-8 stdout (кириллица в выводе на Windows-консоли)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


def cell_diff(v1_path, v2_path, out_path):
    """Открывает v2 как базу (он 'новый'), сравнивает с v1, заливает изменённые
    ячейки. Возвращает счётчики по типам различий. Структура листов — как в v2."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # изменилось
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # добавилось (в v2, нет в v1)
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # удалилось (в v1, нет в v2)

    wb1 = load_workbook(v1_path, data_only=False)
    wb2 = load_workbook(v2_path, data_only=False)
    stats = {"changed": 0, "added": 0, "removed": 0,
             "sheets_only_in_v2": [], "sheets_only_in_v1": []}

    for name in wb2.sheetnames:
        if name not in wb1.sheetnames:
            stats["sheets_only_in_v2"].append(name)
            continue
        ws1, ws2 = wb1[name], wb2[name]
        max_row = max(ws1.max_row, ws2.max_row)
        max_col = max(ws1.max_column, ws2.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v1 = ws1.cell(r, c).value
                v2 = ws2.cell(r, c).value
                if v1 == v2:
                    continue
                target = ws2.cell(r, c)
                if v1 is None and v2 is not None:
                    target.fill = GREEN
                    stats["added"] += 1
                elif v1 is not None and v2 is None:
                    target.fill = RED
                    target.value = f"[DELETED] was: {v1}"
                    stats["removed"] += 1
                else:
                    target.fill = YELLOW
                    stats["changed"] += 1

    for name in wb1.sheetnames:
        if name not in wb2.sheetnames:
            stats["sheets_only_in_v1"].append(name)

    wb2.save(out_path)
    return stats


def formula_diff(v1_path, v2_path):
    """Список расхождений формул: [(sheet, A1, formula_v1, formula_v2), ...].
    Только ячейки, где хотя бы в одном файле есть формула. Ловит и 'формула
    захардкожена в значение' (f1 есть, f2 нет)."""
    from openpyxl import load_workbook
    wb1 = load_workbook(v1_path, data_only=False)
    wb2 = load_workbook(v2_path, data_only=False)
    out = []
    for name in set(wb1.sheetnames) & set(wb2.sheetnames):
        ws1, ws2 = wb1[name], wb2[name]
        for row in ws2.iter_rows():
            for cell in row:
                f2 = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
                f1_val = ws1.cell(cell.row, cell.column).value
                f1 = f1_val if isinstance(f1_val, str) and f1_val.startswith("=") else None
                if f1 != f2 and (f1 or f2):
                    out.append((name, cell.coordinate, f1, f2))
    return out


def find_formula_errors(path):
    """Найти ячейки с ошибками #REF!/#DIV/0!/#NAME? по cached-значениям.
    ВНИМАНИЕ: требует, чтобы файл был открыт/сохранён в Excel (иначе кэша нет
    и формулы вернут None). Возвращает [(sheet, coord, error), ...]."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    errors = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors.append((sheet, cell.coordinate, cell.value))
    return errors


def find_duplicates(path, key, sheet_name=0):
    """Дубликаты по колонке-ключу. Возвращает pandas.DataFrame дублей."""
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet_name)
    dups = df[df.duplicated(subset=[key], keep=False)]
    return dups.sort_values(key)


def _main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 1
    cmd = argv[1]
    if cmd == "celldiff" and len(argv) == 5:
        print(cell_diff(argv[2], argv[3], argv[4]))
    elif cmd == "formuladiff" and len(argv) == 4:
        rows = formula_diff(argv[2], argv[3])
        for sheet, coord, f1, f2 in rows:
            print(f"{sheet}!{coord}:\n  v1: {f1}\n  v2: {f2}")
        print(f"\n{len(rows)} formula diffs")
    elif cmd == "errors" and len(argv) == 3:
        errs = find_formula_errors(argv[2])
        for s, c, e in errs:
            print(f"{s}!{c}: {e}")
        print(f"\n{len(errs)} formula errors")
    elif cmd == "dupes" and len(argv) == 4:
        print(find_duplicates(argv[2], argv[3]).to_string())
    else:
        print(__doc__)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(_main(sys.argv))
