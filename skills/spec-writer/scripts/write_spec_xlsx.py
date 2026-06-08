"""Запись нового листа «Спец. N» в существующий xlsx-реестр объекта <организация>.

Минимально достаточный helper:
- открывает существующий workbook (реестр <объект> / <объект> и т.д.)
- копирует структуру с template-листа (если задан): заголовки, ширины колонок, стили шапки
- записывает позиции из items
- read-back verify

Использование:
    from write_spec_xlsx import write_spec_sheet, verify_spec_sheet

    write_spec_sheet(
        workbook_path="Реестр_<объект>.xlsx",
        sheet_name="Спец. 5",
        items=[
            {"D": "ABC-123", "E": "м", "qty": 1000, "price": 35.50},
            {"D": "XYZ-456", "E": "шт", "qty": 200, "price": 70.00},
        ],
        template_sheet="Спец. 4",
        header_row=1,
        data_start_row=2,
    )

    issues = verify_spec_sheet("Реестр_<объект>.xlsx", "Спец. 5", expected_rows=2)
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional, Iterable
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter


def write_spec_sheet(
    workbook_path: str | Path,
    sheet_name: str,
    items: list[dict],
    template_sheet: Optional[str] = None,
    header_row: int = 1,
    data_start_row: int = 2,
    key_to_column: Optional[dict[str, str]] = None,
) -> dict:
    """Создаёт новый лист и пишет в него позиции.

    Args:
        workbook_path: путь к существующему xlsx (реестру объекта).
        sheet_name: название нового листа (например "Спец. 5").
        items: список словарей-позиций. Ключи маппятся на колонки через `key_to_column`
            или автоматически по заголовкам template_sheet.
        template_sheet: имя листа-эталона для копирования структуры (заголовки + ширины).
        header_row: строка с заголовками в template_sheet.
        data_start_row: строка с которой начинать запись данных в новом листе.
        key_to_column: явный маппинг `{ключ_dict: буква_колонки}`. Если None — берётся из
            заголовков template_sheet (имя заголовка == ключ в item).

    Returns:
        dict с результатом записи (counters + warnings).
    """
    wb_path = Path(workbook_path)
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook не найден: {wb_path}")

    wb = openpyxl.load_workbook(wb_path)

    if sheet_name in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' уже существует. Удали вручную или укажи другое имя.")

    new_ws = wb.create_sheet(sheet_name)
    warnings: list[str] = []

    if template_sheet:
        if template_sheet not in wb.sheetnames:
            raise ValueError(f"Template-лист '{template_sheet}' не найден.")
        tmpl_ws = wb[template_sheet]
        _copy_header(tmpl_ws, new_ws, header_row=header_row)
        _copy_column_widths(tmpl_ws, new_ws)
        if key_to_column is None:
            key_to_column = _infer_key_mapping(tmpl_ws, header_row=header_row)

    if not key_to_column:
        warnings.append("Нет маппинга ключей на колонки — данные не записаны.")
        wb.save(wb_path)
        return {"status": "warning", "rows_written": 0, "warnings": warnings}

    rows_written = 0
    for i, item in enumerate(items):
        row_idx = data_start_row + i
        for key, value in item.items():
            col = key_to_column.get(key)
            if not col:
                warnings.append(f"Ключ '{key}' не в маппинге — пропущен в row {row_idx}.")
                continue
            new_ws[f"{col}{row_idx}"] = value
        rows_written += 1

    wb.save(wb_path)
    return {
        "status": "ok",
        "sheet_name": sheet_name,
        "rows_written": rows_written,
        "warnings": warnings,
    }


def verify_spec_sheet(
    workbook_path: str | Path,
    sheet_name: str,
    expected_rows: Optional[int] = None,
    expected_columns: Optional[Iterable[str]] = None,
    header_row: int = 1,
    data_start_row: int = 2,
) -> list[str]:
    """Read-back verification после write_spec_sheet (§4 Karpathy).

    Returns: список найденных проблем. Пустой список = всё ок.
    """
    issues: list[str] = []
    wb = openpyxl.load_workbook(workbook_path, data_only=False)

    if sheet_name not in wb.sheetnames:
        return [f"Лист '{sheet_name}' не найден в файле"]

    ws = wb[sheet_name]

    if ws.max_row < data_start_row:
        issues.append(f"Лист пустой: max_row={ws.max_row}, ожидалась data_start_row={data_start_row}")

    if expected_rows is not None:
        actual_data_rows = ws.max_row - (data_start_row - 1)
        if actual_data_rows != expected_rows:
            issues.append(f"Ожидали {expected_rows} строк данных, нашли {actual_data_rows}")

    if expected_columns:
        actual_headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        for exp in expected_columns:
            if exp not in actual_headers:
                issues.append(f"Ожидаемый заголовок '{exp}' отсутствует")

    for row in ws.iter_rows(min_row=data_start_row, values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("#") and v in {"#REF!", "#DIV/0!", "#NAME?", "#VALUE!", "#N/A"}:
                issues.append(f"Ошибка формулы в {cell.coordinate}: {v}")

    return issues


def _copy_header(src_ws, dst_ws, header_row: int = 1) -> None:
    """Копирует строку заголовков (значения + стили шрифта/выравнивания/заливки)."""
    for col_idx in range(1, src_ws.max_column + 1):
        src_cell = src_ws.cell(header_row, col_idx)
        dst_cell = dst_ws.cell(header_row, col_idx)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format


def _copy_column_widths(src_ws, dst_ws) -> None:
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = dim.width


def _infer_key_mapping(template_ws, header_row: int = 1) -> dict[str, str]:
    """Из заголовков template-листа собирает маппинг {header_text: column_letter}."""
    mapping: dict[str, str] = {}
    for col_idx in range(1, template_ws.max_column + 1):
        header = template_ws.cell(header_row, col_idx).value
        if header is not None and str(header).strip():
            mapping[str(header).strip()] = get_column_letter(col_idx)
    return mapping


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 3:
        print("Usage: python write_spec_xlsx.py <workbook.xlsx> <sheet_name> [<template_sheet>]")
        sys.exit(1)
    wb_path, sheet = sys.argv[1], sys.argv[2]
    tmpl = sys.argv[3] if len(sys.argv) > 3 else None
    items_demo = [
        {"D": "DEMO-001", "E": "шт", "qty": 1, "price": 100.0},
    ]
    result = write_spec_sheet(wb_path, sheet, items_demo, template_sheet=tmpl)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    issues = verify_spec_sheet(wb_path, sheet, expected_rows=len(items_demo))
    print("Issues:", issues)
