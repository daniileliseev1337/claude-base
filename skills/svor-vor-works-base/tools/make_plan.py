# -*- coding: utf-8 -*-
"""Матчинг блоков на типы базы + план правок (replace/delete/skip) + прочёс подозрительных.

  python make_plan.py --base "<база.xlsx>" --blocks blocks.json --out plan.json
Ключи типов читаются ИЗ базы (лист «База работ», колонка D). FIXUPS ниже — точечные коррекции
ложняков (подстрока работы + неверный тип → верный); специфичные подстроки РАНЬШЕ общих.
После изменения матчера/базы — разобрать секцию «ПОДОЗРИТЕЛЬНЫЕ» глазами до применения.
"""
import openpyxl, json, os, re, argparse
from collections import Counter, defaultdict

# ---------- типы со спецобработкой ----------
DELETE_TYPES = {2, 7}      # нематериальные, крепёж — строку-работу удалить
SKIP_TYPES = {33}          # строительные — не менять
FIXED_TEXT = {34: 'Маркировка кабелей и оборудования',
              36: 'Огнезащитная герметизация проходок'}

# ---------- точечные коррекции ложняков (пополнять по мере находок) ----------
FIXUPS = [
    ('snmp', 1, 27), ('snmp', 6, 27),
    ('кронштейн', 10, 35), ('крепления', 10, 35), ('крепление', 10, 35),
    ('анкера', 10, 7), ('анкера', 35, 7), ('крепежа', 10, 7), ('крепежа', 5, 7),
    ('розетки', 25, 22), ('розетки рабочего места', 1, 22),
    ('щита', 25, 30), ('щита', 5, 30),
    ('лючка', 5, 22),
    ('телефона', 4, 1), ('моста', 4, 1),
    ('источника питания коммутатора', 1, 4),
    ('монитора', 9, 14),
    ('передатчика по витой паре', 9, 1), ('приемника по витой паре', 9, 1),
    ('передатчика и приемника', 9, 1), ('передатчика сигналов по оптическому', 9, 1),
    ('приемника сигналов по оптическому', 9, 1), ('комплекта передачи сигналов', 9, 1),
    ('пульта ptz', 10, 1),
    ('сервера системы палатной', 2, 13), ('сервера системы палатной', 25, 13),
    ('сервера', 14, 1), ('сервера', 2, 1),
    ('маршрутизатора', 28, 1),
    ('свитчера', 15, 1), ('свитчером', 35, 1),
    ('моноблока', 15, 13), ('сенсорного стола', 21, 13),
    ('вентиляционных панелей', 14, 5),
    ('кросса', 22, 26),
    ('фермовой', 22, 35),
    ('разветвителя dmx', 21, 31), ('сплиттера dmx', 29, 31),
    ('адаптера питания', 28, 4), ('ибп', 30, 6),
    ('кабеля hdmi', 25, 9), ('кабеля usb', 25, 9), ('кабеля dmx', 25, 9),
    ('ретрактора', 25, 9), ('удлинителя usb', 25, 19),
    ('разъем', 25, 22), ('разъём', 25, 22),
    ('рукава', 36, 24),
    ('датчика', 23, 28),
    ('винта', 16, 7), ('полки', 16, 5),
    ('заглушки', 31, 7), ('блока питания', 31, 4),
    ('конвертера', 35, 1), ('атс', 2, 1),
    ('рамы подвеса', 21, 35),
    ('оптического передатчика', 25, 1), ('оптического приемника', 25, 1),
    ('оптического приёмника', 25, 1),
    ('угла', 7, 23),
]

# ---------- префиксы существующих работ (для выделения хвоста-объекта) ----------
PREFIXES = ['Проверка комплектности и передача ', 'Установка и настройка ',
            'Настройка и запуск ', 'Монтаж и пусконаладка ', 'Монтаж (оконцевание) ',
            'Огнезащитная герметизация ', 'Пусконаладка ', 'Прокладка ', 'Монтаж ',
            'Установка ', 'Активация ', 'Программирование ', 'Маркировка ',
            'Алмазное бурение ', 'Штробление ']

# ---------- фразовый матчер ----------
SEP = r'[\s\-–—/,\.\(\)«»"]{1,3}'

def stem_word(w):
    if len(w) >= 6: return w[:-2]
    if len(w) >= 4: return w[:-1]
    return w

def key_stems(key):
    words = re.findall(r'[a-zа-яё0-9\-\.\+]+', key.lower())
    out = []
    for w in words:
        if re.search(r'[а-яё]', w):
            s = stem_word(w)
            out.append((s, s != w))
        else:
            out.append((w, False))
    return out

def stem_rx(stem, trimmed):
    if trimmed:
        return re.escape(stem) + r'[а-яё]*'
    if len(stem) <= 3 and not re.search(r'[а-яё]', stem):
        return re.escape(stem) + r'(?![a-z])'
    return re.escape(stem) + (r'[a-z0-9\.\+]*' if not re.search(r'[а-яё]', stem) else '')

def phrase_match(stems, t):
    def rx(seq):
        return r'(?<![a-zа-яё0-9])' + SEP.join(stem_rx(s, tr) for s, tr in seq)
    if re.search(rx(stems), t):
        return True
    return len(stems) == 2 and re.search(rx(stems[::-1]), t) is not None

def load_types(base_path):
    wb = openpyxl.load_workbook(base_path)
    ws = wb['База работ']
    types = []
    for r in range(2, ws.max_row + 1):
        num = ws.cell(r, 1).value
        if num is None:
            continue
        types.append({'num': int(num), 'name': str(ws.cell(r, 2).value or '').strip(),
                      'keywords': [k.strip().lower() for k in str(ws.cell(r, 4).value or '').split(';') if k.strip()],
                      'works_f': str(ws.cell(r, 6).value or '').strip()})
    return types

def match_types(types, text):
    t = text.lower()
    hits = []
    for tp in types:
        best, best_key = 0, ''
        for key in tp['keywords']:
            stems = key_stems(key)
            if stems and phrase_match(stems, t):
                score = sum(len(s) for s, _ in stems)
                if score > best:
                    best, best_key = score, key
        if best:
            hits.append((best, tp['num'], tp['name'], best_key))
    hits.sort(key=lambda h: (-h[0], -h[1]))   # тай-брейк: специфичный (больший №)
    return hits

def apply_fixups(work, tnum):
    wl = work.lower()
    for sub, frm, to in FIXUPS:
        if frm == tnum and sub in wl:
            return to
    return tnum

def enum_phrase(works_f, tail):
    parts = [p.strip() for p in works_f.split(',') if p.strip()]
    if len(parts) == 1:
        phrase = parts[0]
    else:
        phrase = ', '.join(parts[:-1]) + ' и ' + parts[-1][0].lower() + parts[-1][1:]
    phrase = phrase[0].upper() + phrase[1:]
    return (phrase + ' ' + tail).strip()

def tail_of(work):
    for p in PREFIXES:
        if work.startswith(p):
            return work[len(p):].strip()
    return None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--base', required=True)
    ap.add_argument('--blocks', default='blocks.json')
    ap.add_argument('--out', default='plan.json')
    args = ap.parse_args()

    types = load_types(args.base)
    tname = {tp['num']: tp['name'] for tp in types}
    tworks = {tp['num']: tp['works_f'] for tp in types}
    blocks_data = json.load(open(args.blocks, encoding='utf-8'))

    plan, stats, warnings, matched = [], Counter(), [], []
    for sheet in blocks_data:
        fp = {'file': sheet['file'], 'kind': sheet['kind'], 'sheet': sheet['sheet'], 'actions': []}
        for b in sheet['blocks']:
            rows = b['rows']
            n = len(rows)
            wt = str(rows[0]['D'] or '').strip()
            if n == 1 or wt in ('4', 'Наименование работ и затрат') or wt.startswith('Раздел'):
                continue
            k = n // 2
            w_rows = [r['row'] for r in rows[:k]]
            p_rows = rows[k:]
            pos = ' /// '.join(str(r['D'] or '').strip() for r in p_rows)
            # окно позиции 120 симв.: длинные хвосты описаний дают ложняки
            # («в комплекте... сертификат», «по ГОСТ» и т.п.)
            hits = match_types(types, pos[:120] + ' ' + wt)
            if not hits:
                warnings.append(f"{sheet['file'][:40]} row{b['top']}: НЕ ПОКРЫТО: {wt[:50]} /// {pos[:60]}")
                stats['NO_MATCH'] += 1
                continue
            tnum = apply_fixups(wt, hits[0][1])
            matched.append({'work': wt, 'pos': pos[:120], 'type': tnum, 'key': hits[0][3]})
            if tnum in SKIP_TYPES:
                stats['skip'] += 1
                continue
            if tnum in DELETE_TYPES:
                carry = {'A': b['num']}
                for col in ('B', 'C'):
                    if not [r[col] for r in p_rows if r[col] not in (None, '')] and rows[0][col] not in (None, ''):
                        carry[col] = rows[0][col]
                fp['actions'].append({'action': 'delete', 'rows': w_rows, 'carry': carry,
                                      'type': tnum, 'work': wt})
                stats[f'delete_t{tnum}'] += 1
                continue
            if tnum in FIXED_TEXT:
                new_text = FIXED_TEXT[tnum]
            else:
                tail = tail_of(wt)
                if not tail:
                    warnings.append(f"{sheet['file'][:40]} row{b['top']}: префикс не распознан: {wt[:60]}")
                    stats['NO_PREFIX'] += 1
                    continue
                new_text = enum_phrase(tworks[tnum], tail)
            if new_text == wt:
                stats['unchanged'] += 1
                continue
            fp['actions'].append({'action': 'replace', 'rows': w_rows, 'new_text': new_text,
                                  'type': tnum, 'work': wt})
            stats['replace'] += 1
        plan.append(fp)

    json.dump(plan, open(args.out, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    print('=== Статистика:', dict(stats))
    for w in warnings[:30]:
        print('  !', w)

    # прочёс подозрительных — работа не похожа на присвоенный тип (разобрать глазами!)
    print('\n=== ПОДОЗРИТЕЛЬНЫЕ (разобрать глазами до применения):')
    CREP = ('комплект', 'саморез', 'дюбел', 'анкер', 'винт', 'гайк', 'шайб', 'шпильк', 'болт',
            'хомут', 'стяжк', 'скоб', 'держател', 'клипс', 'лент', 'оплетк', 'втулк',
            'наконечник', 'сальник', 'заглушк', 'колпачк', 'уголк', 'пластин', 'профил',
            'траверс', 'материал', 'крепеж', 'крепёж', 'угла', 'углов')
    checks = {2: lambda w: any(s in w for s in ('лиценз', 'сертифик', 'программного обеспеч',
                                                'активац', '—', 'комплектности')),
              7: lambda w: any(s in w for s in CREP),
              25: lambda w: any(s in w for s in ('кабел', 'провод', 'прокладка')),
              24: lambda w: any(s in w for s in ('труб', 'муфт', 'рукав', 'колен', 'ввод')),
              3: lambda w: any(s in w for s in ('шнур', 'патч', 'кабел', 'сборк'))}
    seen = set()
    n_susp = 0
    for m in matched:
        chk = checks.get(m['type'])
        key = (m['work'], m['type'])
        if chk and key not in seen and not chk(m['work'].lower()):
            seen.add(key)
            n_susp += 1
            print(f"  t{m['type']:2d} {m['work'][:60]}  key={m['key'][:30]}")
            print(f"       поз: {m['pos'][:85]}")
    if not n_susp:
        print('  чисто')

if __name__ == '__main__':
    main()
