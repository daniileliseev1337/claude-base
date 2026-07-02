# -*- coding: utf-8 -*-
"""Верификация результата против плана: блоки 1:1, тексты D, переносы №/B/C, формулы H/I.

  python verify_plan.py --plan plan.json --blocks blocks.json --dst-suffix " v2"
0 ошибок = можно на ревьюеров (excel-validator + auditor).
"""
import json, os, re, argparse
import openpyxl

def extract_blocks(ws, header_row):
    a_merges = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1 and mr.max_col == 1 and mr.min_row > header_row and mr.max_row > mr.min_row:
            a_merges[mr.min_row] = mr.max_row
    blocks = []
    r = header_row + 1
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        d = ws.cell(r, 4).value
        if isinstance(a, str) and a.strip().startswith('Раздел'):
            r += 1
            continue
        if r in a_merges:
            top, bot = r, a_merges[r]
            d0 = str(ws.cell(top, 4).value or '').strip()
            if d0 in ('4', 'Наименование работ и затрат') or d0.startswith('Раздел'):
                r = bot + 1
                continue
            blocks.append([{'row': rr, 'A': ws.cell(rr, 1).value, 'B': ws.cell(rr, 2).value,
                            'C': ws.cell(rr, 3).value, 'D': ws.cell(rr, 4).value}
                           for rr in range(top, bot + 1)])
            r = bot + 1
            continue
        if a is not None and d is not None and str(d).strip() not in ('4', 'Наименование работ и затрат'):
            blocks.append([{'row': r, 'A': a, 'B': ws.cell(r, 2).value,
                            'C': ws.cell(r, 3).value, 'D': d}])
        r += 1
    return blocks

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--plan', default='plan.json')
    ap.add_argument('--blocks', default='blocks.json')
    ap.add_argument('--dst-suffix', default=' v2')
    args = ap.parse_args()
    plan = json.load(open(args.plan, encoding='utf-8'))
    blocks_data = json.load(open(args.blocks, encoding='utf-8'))
    bd = {(b['kind'], b['file']): b for b in blocks_data}

    total_err = 0
    for fp in plan:
        src_sheet = bd[(fp['kind'], fp['file'])]
        path = os.path.join(os.path.dirname(src_sheet['path']) + args.dst_suffix, fp['file'])
        if not os.path.exists(path):
            print('SKIP (нет файла)', fp['file'][:60])
            continue
        errs = []
        src_blocks = []
        for b in src_sheet['blocks']:
            wt = str(b['rows'][0]['D'] or '').strip()
            if len(b['rows']) == 1 or wt in ('4', 'Наименование работ и затрат') or wt.startswith('Раздел'):
                continue
            src_blocks.append(b)
        action_by_top = {a['rows'][0]: a for a in fp['actions']}
        ws = openpyxl.load_workbook(path, data_only=False)[fp['sheet']]
        v2 = extract_blocks(ws, src_sheet['header_row'])
        if len(v2) != len(src_blocks):
            errs.append(f'число блоков: v2={len(v2)} vs src={len(src_blocks)}')
        for i, (sb, vb) in enumerate(zip(src_blocks, v2)):
            rows = sb['rows']
            n, k = len(rows), len(rows) // 2
            wt = str(rows[0]['D'] or '').strip()
            act = action_by_top.get(rows[0]['row'])
            if act and act['action'] == 'delete':
                if len(vb) != n - k:
                    errs.append(f'#{i} ({wt[:30]}): delete, размер {len(vb)} != {n-k}')
                    continue
                if str(vb[0]['A']) != str(act['carry'].get('A')):
                    errs.append(f"#{i}: A не перенесён")
                for col in ('B', 'C'):
                    if col in act['carry'] and str(vb[0][col] or '') != str(act['carry'][col] or ''):
                        errs.append(f"#{i}: {col} не перенесён")
                for j, pr in enumerate(rows[k:]):
                    if str(vb[j]['D'] or '').strip() != str(pr['D'] or '').strip():
                        errs.append(f"#{i}: D позиции изменился")
            else:
                if len(vb) != n:
                    errs.append(f'#{i} ({wt[:30]}): размер {len(vb)} != {n}')
                    continue
                expected = act['new_text'] if act else wt
                for j in range(k):
                    if str(vb[j]['D'] or '').strip() != expected:
                        errs.append(f"#{i}: D работы: {str(vb[j]['D'])[:45]!r} != {expected[:45]!r}")
                for j in range(k, n):
                    if str(vb[j]['D'] or '').strip() != str(rows[j]['D'] or '').strip():
                        errs.append(f"#{i}: D позиции изменился")
        if fp['kind'] == 'СВОР':
            bad = 0
            for vb in v2:
                for cell in vb:
                    for col in (8, 9):
                        v = ws.cell(cell['row'], col).value
                        if isinstance(v, str) and v.startswith('='):
                            refs = set(re.findall(r'[FG](\d+)', v))
                            if refs and refs != {str(cell['row'])}:
                                bad += 1
            if bad:
                errs.append(f'формул H/I с чужой строкой: {bad}')
        print(('PASS' if not errs else f'FAIL ({len(errs)})').ljust(10), fp['kind'].ljust(4), fp['file'][:55])
        for e in errs[:8]:
            print('   !', e)
        total_err += len(errs)
    print('\nИТОГО ошибок:', total_err)

if __name__ == '__main__':
    main()
