# -*- coding: utf-8 -*-
"""Извлечение блоков (строки-работы + позиции) из ведомостей СВОР/ВОР.

Блок = вертикальный merge колонки A (№ п/п). ВОР: 2 строки (работа+позиция),
СВОР: до 4 строк (2 работы «до/после» + 2 позиции). Запуск:
  python extract_blocks.py --files "C:\\path\\СВОР\\С работами\\*.xlsx" --files "C:\\path\\ВОР\\...\\*.xlsx" --out blocks.json
Windows: PYTHONIOENCODING=utf-8.
"""
import openpyxl, json, os, glob, argparse

def extract_sheet(ws):
    header_row = None
    for r in range(1, min(ws.max_row, 40) + 1):
        if str(ws.cell(r, 1).value or '').strip().startswith('№'):
            header_row = r
            break
    if header_row is None:
        return None
    a_merges = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1 and mr.max_col == 1 and mr.min_row > header_row and mr.max_row > mr.min_row:
            a_merges[mr.min_row] = mr.max_row
    blocks = []
    section = None
    r = header_row + 1
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        d = ws.cell(r, 4).value
        if isinstance(a, str) and a.strip().startswith('Раздел'):
            section = a.strip()
            r += 1
            continue
        if r in a_merges:
            top, bot = r, a_merges[r]
            rows = [{'row': rr, 'B': ws.cell(rr, 2).value, 'C': ws.cell(rr, 3).value,
                     'D': ws.cell(rr, 4).value, 'E': ws.cell(rr, 5).value}
                    for rr in range(top, bot + 1)]
            blocks.append({'num': a, 'section': section, 'top': top, 'bot': bot, 'rows': rows})
            r = bot + 1
            continue
        if a is not None and d is not None:
            blocks.append({'num': a, 'section': section, 'top': r, 'bot': r, 'no_merge': True,
                           'rows': [{'row': r, 'B': ws.cell(r, 2).value, 'C': ws.cell(r, 3).value,
                                     'D': d, 'E': ws.cell(r, 5).value}]})
        r += 1
    return header_row, blocks

def clean(o):
    if isinstance(o, dict):
        return {k: clean(v) for k, v in o.items()}
    if isinstance(o, list):
        return [clean(x) for x in o]
    return o if o is None or isinstance(o, (int, float, str)) else str(o)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--files', action='append', required=True, help='glob путей xlsx (можно несколько раз)')
    ap.add_argument('--out', default='blocks.json')
    args = ap.parse_args()
    paths = []
    for g in args.files:
        paths += sorted(glob.glob(g))
    result = []
    for path in paths:
        kind = 'СВОР' if 'СВОР' in path.upper() else 'ВОР'
        wb = openpyxl.load_workbook(path, data_only=False)
        for ws in wb.worksheets:
            res = extract_sheet(ws)
            if res is None:
                continue
            header_row, blocks = res
            result.append({'file': os.path.basename(path), 'path': path, 'kind': kind,
                           'sheet': ws.title, 'header_row': header_row,
                           'n_blocks': len(blocks), 'blocks': blocks})
    json.dump(clean(result), open(args.out, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    print('sheets:', len(result), 'blocks:', sum(x['n_blocks'] for x in result), '->', args.out)

if __name__ == '__main__':
    main()
